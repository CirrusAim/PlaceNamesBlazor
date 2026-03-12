"""Create sample batch import Excel (sheet 'Fields') with 5 rows. Run: python create_sample_excel.py"""
try:
    from openpyxl import Workbook
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Fields"

# Headers (required + optional batch fields; optional columns can be omitted in Excel and import still works)
headers = [
    "Fylke", "Stempletype", "Poststed", "Kommune", "Stempeltekst", "Første kjente", "Siste kjente", "Kommentar", "Bilde",
    "Stempeldiameter", "Bokstavhøyde", "Andre mål", "Stempelfarge", "Tapsmelding", "Reparasjoner", "Dato avtrykk i PM",
    "Dato fra gravør", "Dato fra intendantur", "Dato fra overordnet", "Dato for innlevering", "Dato innlevert intendantur",
]
for c, h in enumerate(headers, 1):
    ws.cell(row=1, column=c, value=h)

# 5 sample rows with full data for every field (Fylke/Stempletype match full_schema.sql seeds)
bilde_path = r"C:\Users\cirru\Downloads\wiVRVa42.jpg"
# Columns: Fylke, Stempletype, Poststed, Kommune, Stempeltekst, Første kjente, Siste kjente, Kommentar, Bilde,
#          Stempeldiameter, Bokstavhøyde, Andre mål, Stempelfarge, Tapsmelding, Reparasjoner, Dato avtrykk i PM,
#          Dato fra gravør, Dato fra intendantur, Dato fra overordnet, Dato for innlevering, Dato innlevert intendantur
rows = [
    [
        "Akershus", "I", "Ski", "Nordre Follo", "Ski",
        "2020-01-15", "2024-06-01",
        "Eksempelstempel 1, nordre Follo.",
        bilde_path,
        25.5, 3.2, "H=10mm, B=8mm",
        "Svart", "—", "Ingen",
        "2020-02-01", "2020-01-01", "2020-01-10", "2020-01-15", "2020-01-20", "2020-01-25",
    ],
    [
        "Bergen", "SL", "Bergen", "Bergen", "Bergen sentrum",
        "2019-05-10", "2023-12-31",
        "Eksempelstempel 2, Bergen by.",
        bilde_path,
        24.0, 3.0, "Diameter 24 mm",
        "Blå", "Tapsmelding 12.03.2020", "Liten reparasjon 2021",
        "2019-06-15", "2019-05-20", "2019-06-01", "2019-06-05", "2019-06-10", "2019-06-12",
    ],
    [
        "Rogaland", "SA", "Stavanger", "Stavanger", "Stavanger",
        "2018-03-01", "2022-08-15",
        "Eksempelstempel 3, Stavanger.",
        bilde_path,
        26.0, 3.5, "Høyde 12 mm, bredde 8 mm",
        "Svart", "—", "—",
        "2018-04-01", "2018-03-15", "2018-03-20", "2018-03-25", "2018-04-01", "2018-04-05",
    ],
    [
        "Sør-Trøndelag", "III", "Trondheim", "Trondheim", "Trondheim",
        "2021-02-20", "2024-01-10",
        "Eksempelstempel 4, Trondheim sentrum.",
        bilde_path,
        25.0, 3.2, "Rundt 25 mm",
        "Rød", "Tapsmelding 2022", "Skift fjær 2023",
        "2021-03-01", "2021-02-25", "2021-03-05", "2021-03-08", "2021-03-15", "2021-03-20",
    ],
    [
        "Østfold", "I20", "Moss", "Moss", "Moss",
        "2020-11-01", "2023-09-30",
        "Eksempelstempel 5, Moss postkontor.",
        bilde_path,
        23.5, 2.8, "Oval, 23×18 mm",
        "Svart", "—", "Ingen reparasjoner",
        "2020-12-01", "2020-11-15", "2020-11-20", "2020-11-25", "2020-12-01", "2020-12-05",
    ],
]
for r, row in enumerate(rows, 2):
    for c, val in enumerate(row, 1):
        ws.cell(row=r, column=c, value=val if val else None)

out = "SampleBatchImport.xlsx"
wb.save(out)
print(f"Created {out}")
